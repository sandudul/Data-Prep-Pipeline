import csv
import json
from typing import Any, Dict, List
from .core import DataReader, DataType

try:
    import openpyxl
except ImportError:
    openpyxl = None

class CSVReader(DataReader):
    """Reads data from a CSV file."""
    
    def __init__(self, filepath: str, delimiter: str = ','):
        self.filepath = filepath
        self.delimiter = delimiter
        
    def read(self) -> DataType:
        data: DataType = []
        with open(self.filepath, mode='r', encoding='utf-8') as f:
            reader = csv.DictReader(f, delimiter=self.delimiter)
            for row in reader:
                data.append(dict(row))
        return data


class JSONReader(DataReader):
    """Reads data from a JSON file (expecting a list of objects)."""
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        
    def read(self) -> DataType:
        with open(self.filepath, mode='r', encoding='utf-8') as f:
            data = json.load(f)
            
        if not isinstance(data, list):
            raise ValueError("JSON file must contain a list of objects.")
            
        # Ensure it's a list of dicts
        for item in data:
            if not isinstance(item, dict):
                raise ValueError("JSON list items must be objects/dictionaries.")
                
        return data

class ExcelReader(DataReader):
    """Reads data from an Excel file (.xlsx)."""
    
    def __init__(self, filepath: str, sheet_name: str = None):
        self.filepath = filepath
        self.sheet_name = sheet_name
        
    def read(self) -> DataType:
        if openpyxl is None:
            raise ImportError("The 'openpyxl' package is required to read Excel files. Install it using 'pip install openpyxl'.")
            
        workbook = openpyxl.load_workbook(self.filepath, data_only=True)
        if self.sheet_name:
            if self.sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{self.sheet_name}' not found in the workbook.")
            sheet = workbook[self.sheet_name]
        else:
            sheet = workbook.active
            
        data: DataType = []
        headers = []
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx == 0:
                # Use first row as headers, handle potential None headers
                headers = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(row)]
            else:
                row_dict = {headers[i]: cell for i, cell in enumerate(row) if i < len(headers)}
                data.append(row_dict)
                
        return data
